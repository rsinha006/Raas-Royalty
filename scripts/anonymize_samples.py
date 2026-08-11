#!/usr/bin/env python3
"""Generate anonymized fixtures from samples/.

The files in samples/ are the real, unedited artifacts from past years. They
carry names, phone numbers and contact details for ~250 real people and must
never reach this repository. This script rewrites them into fixtures/ with every
person and phone number replaced, and every structural quirk left intact —
because the quirks are what the importer has to survive.

Run:  python3 scripts/anonymize_samples.py

Deterministic: the same samples/ always produce byte-identical fixtures, so
regenerating doesn't churn the diff. The real -> fake mapping is derived at
runtime and never written to disk.

What is preserved on purpose (see docs/sample-data-analysis.md):

  * merged-range geometry, sheet order, sheet names incl. trailing spaces
  * formulas, side tables in columns G-K, empty sheets
  * name collisions — two real people sharing a name still share one afterwards
  * misspelling pairs — 'Rhea Mehta' / 'Rhea Metha' become two fake spellings
    that are still one edit apart, so fuzzy-match tests keep their teeth
  * trailing whitespace and the '*' / '**' food-restriction suffix on names
  * all five phone formats, including the Unicode-direction-mark one
  * the dancer with no phone number at all
  * surname frequency — 'Patel' is common in the source and stays common

What is replaced: person names (in structured columns AND inside free text),
and phone numbers. Team names, venue names and street addresses are left alone;
they are public facts about organizations and buildings, and the location-alias
edge case is worthless without them.
"""

from __future__ import annotations

import posixpath
import random
import re
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SAMPLES = ROOT / "samples"
FIXTURES = ROOT / "fixtures"

SEED = 20260805
FIXED_TIME = datetime(2026, 1, 1)

# Where a cell IS a person's name rather than free text that might mention one.
# Deliberately an explicit whitelist: harvesting by shape alone pulls in
# 'Rehearsal Schedule' and 'Report to Lobby' as people, and once a word like
# "Schedule" is in the surname map it corrupts every sheet title in the set.
#
# Column B of the four day grids is the people block; see the analysis. The two
# RRXIV files keep their names in one sheet each.
DAY_SHEETS = ("Thursday Schedule", "Friday Schedule", "Saturday Schedule ", "Sunday Schedule")
FULL_NAME_SHEETS = ("Royalty Members", "Deckard Bussing Contacts")

# Two or three capitalized tokens, nothing else. Anchors both ends so
# 'Report to Lobby' (lowercase 'to') and 'BREAK (Aud Staff)' don't qualify.
PERSON_RE = re.compile(r"[A-Z][a-zA-Z'’.-]+(?:\s+[A-Z][a-zA-Z'’.-]+){1,2}")

# Words that pass PERSON_RE but are not people. Checked per token, lowercased.
#
# This list is what makes the free-text sweep safe to run broadly. The sweep
# treats any Title-Case two-or-three-word run as a person unless a token appears
# here — so the failure mode is over-scrubbing an activity name (cosmetic) rather
# than leaking a real one (not cosmetic). Curated by auditing every candidate the
# sweep produced against samples/; contains no personal names, which is why it is
# safe to commit.
NOT_A_NAME = {
    # structure, roles, documents
    "schedule", "rehearsal", "practice", "report", "lobby", "load", "vans", "bus",
    "team", "teams", "board", "members", "liaison", "liaisons", "liasons", "name",
    "position", "royalty", "raas", "break", "stage", "room", "rooms", "hotel",
    "auditorium", "center", "convention", "primary", "office", "number", "total",
    "size", "rig", "upstage", "midstage", "downstage", "director", "directors",
    "captain", "captains", "judge", "judges", "judging", "photo", "show", "order",
    "prop", "props", "drivers", "driving", "airport", "arrivals", "departure",
    "videographer", "freshreps", "hospitality", "logistics", "creative", "funding",
    # days and time
    "friday", "saturday", "sunday", "thursday", "monday", "tuesday", "wednesday",
    "time", "times", "specific", "timer", "starts", "return", "leave", "arrival",
    "arrives", "landing", "day", "night", "morning", "wakey", "wakeup",
    # places
    "aud", "balcony", "courtyard", "marriott", "bloomington", "downtown", "indy",
    "discovery", "parkway", "east", "west", "north", "south", "street", "avenue",
    "college", "walnut", "kinser", "pike", "fieldhouse", "field", "house", "studio",
    "dance", "springhill", "suites", "atrium", "sports", "bar", "bars", "kilroy's",
    "hoosier", "hoosher", "hungama", "bhangra", "nova", "novaktv", "deckard",
    "bill", "garrett", "garret", "campus", "wellness", "mixer", "philanthropy",
    "philanthrophy", "table", "pregame", "venue", "addresses", "important",
    "location", "transit", "great", "registration", "viewing", "dressing",
    "competition", "intramural", "monroe", "county", "walking", "reach",
    # activities and objects
    "after", "party", "car", "cars", "prius", "minibus", "suv", "personal",
    "announce", "results", "pizza", "pizzas", "aver's", "speaking", "ballet",
    "folkorico", "tinikiling", "banner", "height", "width", "bid", "lineup",
    "video", "videos", "intro", "introduction", "charades", "left", "right",
    "clean", "cleanup", "up", "breakfast", "lunch", "dinner", "snack", "snacks",
    "cup", "stacking", "key", "deliver", "dietary", "restrictions", "thank",
    "you", "pickup", "drive", "back", "home", "trophies", "driver", "contact",
    "drop", "off", "eat", "be", "egg", "relay", "faa", "pick", "picked", "find",
    "get", "changed", "ready", "gluten", "free", "going", "out", "gold", "curtain",
    "outhh", "help", "illini", "basket", "illinois", "line", "plot", "lounge",
    "hang", "toss", "lasso", "national", "anthems", "screen", "image", "tech",
    "notes", "questions", "materials", "pack", "computer", "poppy", "seed",
    "sesame", "allergy", "tree", "nut", "think", "beef", "call", "on-call",
    "vehicles", "used", "options", "transportation", "weekend", "boomerang",
    "loading", "rosters", "staff", "michigan", "wolveraas", "setup", "handling",
    "coffee", "more", "need", "stuff", "from", "and", "with", "the", "sober",
    "monitoring", "dad", "mom", "fam", "phone", "top", "keys",
    # Generic English that can pair with a real name in a Title-Case run and
    # get harvested as one ('Friday Post Mixer Practice' -> 'Post').
    "post", "pre", "mid", "first", "last", "next", "new", "old", "main", "full",
    "half", "all", "each", "every", "some", "most", "best", "big", "small",
    "long", "short", "high", "low", "open", "close", "start", "end", "stop",
    "take", "bring", "put", "keep", "make", "give", "send", "meet", "wait",
    "watch", "check", "set", "run", "walk", "sit", "stand", "look", "see",
    "tell", "ask", "use", "work", "play", "move", "turn", "hand", "over",
    # Function words. Critical: substitution is case-insensitive, so letting
    # 'In' through (from 'Arrival In Bloomington') rewrites every lowercase
    # "in" in the workbook.
    "in", "on", "at", "to", "of", "by", "up", "no", "yes", "none", "n/a", "or",
    "but", "for", "from", "with", "without", "before", "after", "into", "onto",
    "out", "off", "back", "down", "under", "again", "then", "than", "this",
    "that", "these", "those", "here", "there", "when", "where", "what", "which",
    "who", "how", "why", "will", "would", "can", "could", "should", "must",
    "may", "might", "does", "did", "is", "are", "was", "were", "been", "being",
    "have", "has", "had", "not", "only", "also", "just", "very", "more", "less",
    "much", "many", "few", "own", "same", "such", "still", "even", "ever",
    "never", "always", "once", "twice", "both", "either", "neither",
    # Roster and event vocabulary that pairs with a real name.
    "senior", "advisor", "advisors", "shirt", "t-shirt", "tshirt", "beef",
    "chicken", "meat", "vegetarian", "vegan", "veg", "non-veg", "allergy",
    "allergies", "nuts", "dairy", "jain", "halal", "kosher", "food", "basket",
    "baskets", "quantity", "count", "counts", "people", "person", "contact",
    "information", "roster", "full", "sheet", "sheets", "tab", "tabs",
    # Shouted nouns that share a cell with a shouted name, and so ride in on
    # tier 3 ('RUNNER\nSEATING', 'BACKSTAGE REGULATOR POST-PROP ROOM').
    "backstage", "regulator", "breakdown", "greet", "them", "seating",
    "running", "runner", "strategy", "gtfo", "dressing", "logs", "meeting",
    # Teams, across both years. These are organizations, not people, and the
    # roster's "team is the sheet name" edge case dies without them. They reach
    # column B of the day grids as group labels, where they otherwise pass for
    # two-token personal names.
    "msu", "umich", "utd", "unc", "gt", "illini", "umd", "uva", "ucsd", "ucla",
    "vt", "gw", "purdue", "chicago", "duke", "raasparty", "taraas", "hooraas",
    "entouraas", "wolveraas", "ruckus", "dhamaal", "bns", "taar", "heel",
    "ramblin", "virginia", "michigan", "maryland", "carolina", "georgia",
    "texas", "dallas", "state", "tech", "iu", "iufs", "ras", "rep", "reps",
}

FAKE_FIRST = """
Aarav Aditi Advait Ahana Akhila Amara Anaya Aneesh Anika Arnav Aruna Ashwin
Avani Bhavna Chetan Chirag Darsh Deepa Dhara Divit Esha Falguni Gaurav Girish
Hansa Harini Hemal Indira Ishan Jaya Jignesh Kalpana Kanan Kavir Keshav Kiran
Lalita Latha Mahesh Maitri Manas Meera Mihika Mohit Naina Namrata Nilay Nirav
Ojas Omkar Pallavi Parag Pooja Prisha Purvi Rachit Radha Rajan Rekha Rishi
Ritvik Rohit Ruchi Sagar Sahana Samir Sanya Sarika Shaan Shalini Sharda Shilpa
Sneha Sohan Sumit Suresh Swara Tanvi Tarak Tejal Uday Ujwal Umang Urmila Vaidehi
Varun Vedika Vihaan Vinita Viraj Yamini Yogesh Zara Neel Ira Dia Kabir Nita Ravi
Bela Chaya Devan Gita Hari Isha Jai Kali Lakshmi Mira Nikhil Priya Rani Sita
Tara Uma Vikram Anya Bhargav Charu Dinesh Ekta Farhan Gopal Heena Imran Janvi
Karan Leela Manoj Nandita Palak Qadir Rupal Saanvi Tanmay Usha Vandana Wasim
Yash Zoya Aisha Bharat Chandan Damini Eshan Ganesh Hitesh Ila Jatin Komal Lavanya
Mahika Nakul Ojasvi Prakash Rhea Sameera Trisha Vivaan Yusuf Ananya Bhaskar
Chitra Devika Elina Gaurika Hemant Ishita Janaki Kunal Lila Madhav Neha Om
Parvati Rajesh Saloni Tanish Vaani Zain Amit Bina Chandra Dhruv Eshwar Geeta
""".split()

FAKE_LAST = """
Acharya Agarwal Ahuja Anand Apte Bajaj Balan Banerjee Bhatt Bose Chandra Chauhan
Chopra Dalal Dave Deshmukh Dhawan Dixit Dutta Gandhi Ghosh Gill Gokhale Grover
Gupta Iyer Jain Joshi Kadam Kamath Kapoor Kaul Khanna Kohli Kulkarni Kumar Lal
Madan Malhotra Mallick Mehra Menon Mishra Mittal Nadkarni Nair Nanda Narang
Nayak Oberoi Pandit Parekh Pillai Prasad Puri Raman Rao Rastogi Sarin Saxena
Sekhar Sengupta Seth Shenoy Sinha Sodhi Subramanian Suri Talwar Tandon Thakur
Tiwari Trivedi Vaidya Varma Venkatesan Verma Vohra Wadhwa Zutshi Bhalla Chadha
Dhingra Fernandes Hegde Kale Lamba Mahajan Nigam Ohri Pathak Rege Sahni Tejwani
Uppal Vasudevan Yadav Bakshi Chawla Damle Gadre Hooda Jaggi Kohli Loyal Mangal
""".split()

# Applied to a fake surname to make a second spelling of it. Each returns a
# string one or two edits away, mirroring the kinds of typo in the real data
# (transposition, dropped letter, doubled letter, vowel swap).
def _transpose(s: str) -> str:
    i = max(1, len(s) // 2)
    if i + 1 >= len(s):
        return s + s[-1]
    return s[:i] + s[i + 1] + s[i] + s[i + 2 :]


def _drop(s: str) -> str:
    i = max(1, len(s) // 2)
    return s[:i] + s[i + 1 :] if len(s) > 3 else s + s[-1]


def _double(s: str) -> str:
    i = max(1, len(s) // 2)
    return s[:i] + s[i] + s[i:]


def _vowel(s: str) -> str:
    for i, ch in enumerate(s):
        if i and ch.lower() in "aeiou":
            repl = {"a": "u", "e": "a", "i": "e", "o": "a", "u": "o"}[ch.lower()]
            return s[:i] + (repl.upper() if ch.isupper() else repl) + s[i + 1 :]
    return _drop(s)


TYPO_FNS = [_transpose, _drop, _double, _vowel]

# The curated lists above are not big enough on their own: the samples carry
# ~310 distinct first names and ~230 surnames, and falling back to numbered
# names ('Mehar2') both looks synthetic and leaves the real name readable as a
# prefix. Extend both pools by syllable so there is always a clean name to hand.
_F_A = "Ka Ma Ra Sa Ni Ve Ta Ja Ha Pa Da Ba Ga La Na Va Cha Bha Dha Sha Kri Pri Tri Anu Ami Ind Yash Ish Ojas Uma".split()
_F_B = "ran vin dev esh ith ika ana ira ela oor ari eev ash ush avi ita nya reya lini mesh".split()
_L_A = "Bhar Chan Dham Gaur Jha Kant Lodh Mahi Nand Pand Rath San Tha Vach Yadu Amb Bir Chit Dev Gan".split()
_L_B = "wala kar nia dev raj esh ani veda pathi shekar bhai wan sena murti pal deep raman jani vati".split()

FAKE_FIRST += [a + b for a in _F_A for b in _F_B]
FAKE_LAST += [a + b for a in _L_A for b in _L_B]

# Marker suffixes that must survive on the fake name. In the real roster these
# denote food restrictions, not captains — see docs/decisions.md.
MARKER_RE = re.compile(r"^(?P<lead>\s*)(?P<core>.*?)(?P<mark>\s*\*{1,2})?(?P<trail>\s*)$", re.S)

DIRECTION_MARKS = "".join(chr(c) for c in range(0x202A, 0x202F))

# One pattern for every phone format in the samples, capturing the punctuation
# so the fake number wears the original's formatting. The set is bigger than it
# looks: bare 10-digit, hyphenated, dotted, space-separated, '(812) 335-8000',
# and '(925)-430-8287' — parens AND a hyphen, which a paren-specific pattern
# misses. Enumerating formats one at a time is how the last one got through.
PHONE_RE = re.compile(
    r"(?<!\d)(\(?)(\d{3})(\)?[-.\s]{0,2})(\d{3})([-.\s]?)(\d{4})(?!\d)"
)


def strip_marks(s: str) -> str:
    return "".join(c for c in s if c not in DIRECTION_MARKS)


def split_name(raw: str):
    """Split 'Aaryan *' into ('', 'Aaryan', ' *', '') so markers survive."""
    m = MARKER_RE.match(raw)
    return m.group("lead"), m.group("core"), m.group("mark") or "", m.group("trail")


def edit_distance(a: str, b: str) -> int:
    if abs(len(a) - len(b)) > 2:
        return 99
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def match_case(fake: str, real: str) -> str:
    """Make the fake token wear the real token's capitalization."""
    if real.isupper() and len(real) > 1:
        return fake.upper()
    if real.islower():
        return fake.lower()
    return fake


class Anonymizer:
    def __init__(self):
        self.rng = random.Random(SEED)
        self.first_map: dict[str, str] = {}
        self.last_map: dict[str, str] = {}
        self.phone_map: dict[str, str] = {}
        self._first_pool: list[str] = []
        self._last_pool: list[str] = []
        self._token_re: re.Pattern | None = None

    # ---------- inventory ----------

    @staticmethod
    def _looks_like_person(s: str) -> bool:
        s = s.strip()
        if not PERSON_RE.fullmatch(s):
            return False
        toks = s.split()
        if any(t.isupper() and len(t) > 1 for t in toks):
            return False  # 'MSU RaasParty', 'UTD TARAAS' — group labels
        return not any(t.lower().strip(".'’-") in NOT_A_NAME for t in toks)

    def harvest(self, books: dict[str, openpyxl.Workbook]):
        firsts, lasts, full = set(), set(), set()

        for fname, wb in books.items():
            # The roster: one sheet per team, First Name in A, Last Name in B.
            if fname.startswith("Team Contact Information"):
                for ws in wb.worksheets[:8]:
                    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                        a, b = row[0], row[1]
                        if isinstance(a, str) and a.strip():
                            firsts.add(split_name(a)[1])
                        if isinstance(b, str) and b.strip():
                            lasts.add(split_name(b)[1])

            # The day grids: column B is one row per person.
            for title in DAY_SHEETS:
                if title in wb.sheetnames:
                    ws = wb[title]
                    for r in range(1, ws.max_row + 1):
                        v = ws.cell(r, 2).value
                        if isinstance(v, str) and self._looks_like_person(v):
                            full.add(v.strip())

            # Name/value sheets in the RRXIV files.
            for title in FULL_NAME_SHEETS:
                if title in wb.sheetnames:
                    for row in wb[title].iter_rows(values_only=True):
                        for v in row:
                            if isinstance(v, str) and self._looks_like_person(v):
                                full.add(v.strip())

        # Tier 2: sweep every cell for people the structured columns miss.
        # The airport and driving tabs name individual dancers, videographers
        # and vendor contacts inline in free text, and those people appear in no
        # roster column anywhere. Missing one is a leak, so this errs toward
        # over-matching and leans on NOT_A_NAME to hold the line.
        for wb in books.values():
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if not isinstance(c.value, str):
                            continue
                        for m in PERSON_RE.finditer(c.value):
                            s = m.group(0).strip()
                            toks = s.split()
                            if any(t.isupper() and len(t) > 1 for t in toks):
                                continue  # ALL-CAPS free text, not a name
                            if any(len(t.strip(".'’-")) < 4 for t in toks):
                                continue  # short tokens are function words far
                                          # more often than names, and a wrong
                                          # one rewrites the whole workbook
                            bad = [t for t in toks
                                   if t.lower().strip(".'’-") in NOT_A_NAME]
                            if not bad:
                                full.add(s)
                            elif len(toks) == 2 and len(bad) == 1:
                                # 'Find Alayna' / 'Babita Speaking' — one half is
                                # a known word, so the other half is a person.
                                solo = next(t for t in toks if t not in bad)
                                if solo.isalpha() and len(solo) >= 4:
                                    firsts.add(solo)

        for s in full:
            parts = s.split()
            firsts.add(parts[0])
            lasts.add(" ".join(parts[1:]))

        firsts = {f for f in firsts if len(f) >= 2 and f[0].isalpha()
                  and f.lower() not in NOT_A_NAME}
        lasts = {l for l in lasts if len(l) >= 2 and l[0].isalpha()
                 and l.lower() not in NOT_A_NAME}

        # Tier 3: ALL-CAPS names. Tiers 1 and 2 skip anything shouted, because
        # the grids are full of shouted prose ('GO TO SLEEP') that reads as a
        # name. But real names get shouted too — the orphan cell at
        # Saturday!Q67 is one, for a person who has no row anywhere. Rescue
        # them by requiring a partner token that is already a known name.
        known = {n.lower() for n in firsts} | {n.lower() for n in lasts}
        for wb in books.values():
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        if not isinstance(c.value, str):
                            continue
                        for run in re.finditer(r"[A-Z][A-Z'’.-]{3,}(?:\s+[A-Z][A-Z'’.-]{3,})+",
                                               c.value):
                            toks = run.group(0).split()
                            if not any(t.lower() in known for t in toks):
                                continue
                            for t in toks:
                                if t.isalpha() and t.lower() not in NOT_A_NAME:
                                    firsts.add(t.capitalize())

        self._assign(sorted(firsts), sorted(lasts))

    def _assign(self, firsts: list[str], lasts: list[str]):
        """Map real -> fake 1:1, keeping near-duplicate surnames near-duplicate."""
        # Drop any fake name that collides with a real one. Both pools are
        # drawn from the same naming traditions as the source data, so overlap
        # is guaranteed — and an overlapping fake is worse than useless: it
        # leaves a real name sitting in the fixture, and makes the leak check
        # unable to tell that name from an intentional one.
        real = {n.lower() for n in firsts} | {n.lower() for n in lasts}
        pool_f = [n for n in FAKE_FIRST if n.lower() not in real]
        pool_l = [n for n in FAKE_LAST if n.lower() not in real]
        if not pool_f or not pool_l:
            raise SystemExit("fake name pools exhausted by collisions with real names")
        self.rng.shuffle(pool_f)
        self.rng.shuffle(pool_l)

        # Cluster surnames that are within 2 edits of each other. The first
        # member gets a clean fake; the rest get typo'd variants of it, so the
        # "same human, two spellings" case survives anonymization.
        clusters: list[list[str]] = []
        for name in lasts:
            for cl in clusters:
                if edit_distance(name.lower(), cl[0].lower()) <= 2:
                    cl.append(name)
                    break
            else:
                clusters.append([name])

        used: set[str] = set()
        cursor = {"f": 0, "l": 0}

        def take(pool: list[str], key: str) -> str:
            """Next unused name from the pool. Never invents a numbered one."""
            while cursor[key] < len(pool):
                cand = pool[cursor[key]]
                cursor[key] += 1
                if cand.lower() not in used:
                    used.add(cand.lower())
                    return cand
            raise SystemExit(
                f"fake {key} pool exhausted — widen FAKE_FIRST / FAKE_LAST"
            )

        for name in firsts:
            self.first_map[name.lower()] = take(pool_f, "f")

        for cl in clusters:
            base = take(pool_l, "l")
            self.last_map[cl[0].lower()] = base
            for j, variant in enumerate(cl[1:]):
                # A near-duplicate of the base, so 'same human, two spellings'
                # survives. Must dodge real names too: a typo chain that lands
                # on one would put it straight back into the fixture.
                v = TYPO_FNS[j % len(TYPO_FNS)](base)
                tries = 0
                while v.lower() in used or v.lower() in real or v == base:
                    v = TYPO_FNS[(j + tries) % len(TYPO_FNS)](v)
                    tries += 1
                    if tries > 8:
                        v = f"{base}{'x' * (j + 1)}"
                        break
                used.add(v.lower())
                self.last_map[variant.lower()] = v

        # Longest first so 'Nitya Shah' is consumed before 'Nitya'.
        tokens = sorted(
            set(list(self.first_map) + list(self.last_map)), key=len, reverse=True
        )
        self._token_re = re.compile(
            r"(?<![A-Za-z])(" + "|".join(re.escape(t) for t in tokens) + r")(?![A-Za-z])",
            re.IGNORECASE,
        )

    # ---------- replacement ----------

    def _sub_token(self, m: re.Match) -> str:
        real = m.group(1)
        key = real.lower()
        fake = self.first_map.get(key) or self.last_map.get(key)
        return match_case(fake, real) if fake else real

    def fake_phone(self, digits: str) -> str:
        if digits not in self.phone_map:
            r = random.Random(SEED + int(digits))
            # 555 is not an assignable NANP area code, so nothing here can dial.
            self.phone_map[digits] = "555" + "".join(str(r.randrange(10)) for _ in range(7))
        return self.phone_map[digits]

    def scrub_text(self, s: str) -> str:
        marks = [(i, c) for i, c in enumerate(s) if c in DIRECTION_MARKS]
        body = strip_marks(s)

        def repl(m):
            f = self.fake_phone(m.group(2) + m.group(4) + m.group(6))
            return f"{m.group(1)}{f[0:3]}{m.group(3)}{f[3:6]}{m.group(5)}{f[6:]}"

        body = PHONE_RE.sub(repl, body)

        if self._token_re:
            body = self._token_re.sub(self._sub_token, body)

        for i, c in marks:  # put the invisible direction marks back where they were
            body = body[:i] + c + body[i:]
        return body

    def scrub_cell(self, v):
        if isinstance(v, str):
            if v.startswith("="):
                return v  # formula: no PII, and rewriting it would break refs
            return self.scrub_text(v)
        if isinstance(v, float) and 1e9 <= v < 1e10 and v == int(v):
            return float(self.fake_phone(str(int(v))))
        if isinstance(v, int) and 1e9 <= v < 1e10:
            return int(self.fake_phone(str(v)))
        return v


# --------------------------------------------------------------------------
# Repack: make openpyxl's output readable by the app, and byte-stable
# --------------------------------------------------------------------------
#
# openpyxl writes a valid .xlsx that Excel opens happily and that the app's
# reader cannot open at all. Two differences from what Excel itself writes, and
# exceljs (server/sync/parse.js) trips on both — it dereferences the part it
# looked for without checking, so the failure surfaces as a TypeError from
# inside a spreadsheet library rather than as anything about the file:
#
#   * relationship targets are absolute (`/xl/tables/table1.xml`) where Excel
#     writes them relative to the part that owns them (`../tables/table1.xml`).
#     exceljs keys its parsed parts by the relative form.
#   * comments live at `xl/comments/comment1.xml` with the VML at
#     `commentsDrawing1.vml`; exceljs only recognises `xl/comments1.xml` and
#     `xl/drawings/vmlDrawing1.vml`.
#
# So the fixtures are rewritten into the layout Excel uses. Nothing about the
# content changes — same parts, same bytes inside them, same order — which is
# what keeps the structure checks in verify_fixtures.py meaningful.
#
# The same pass fixes the timestamps. openpyxl stamps each zip entry with the
# time it ran, so two identical runs produced two different files and the
# docstring's "byte-identical" was not true. Every entry is written at
# FIXED_TIME instead, and now it is.

REL_ELEMENT = re.compile(rb"<Relationship\b[^>]*?/>")
REL_TARGET = re.compile(rb'Target="([^"]*)"')
COMMENT_PART = re.compile(r"^xl/comments/comment(\d+)\.xml$")
COMMENT_VML_PART = re.compile(r"^xl/drawings/commentsDrawing(\d+)\.vml$")
MODIFIED = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")


def _rename_map(names: list[str]) -> dict[str, str]:
    """openpyxl's part paths → the ones Excel writes, for the parts that differ."""
    renames = {}
    for name in names:
        m = COMMENT_PART.match(name)
        if m:
            renames[name] = f"xl/comments{m.group(1)}.xml"
            continue
        m = COMMENT_VML_PART.match(name)
        if m:
            renames[name] = f"xl/drawings/vmlDrawing{m.group(1)}.vml"
    return renames


def _rewrite_rels(data: bytes, rels_path: str, renames: dict[str, str]) -> bytes:
    """Resolve every target to its part, apply the renames, re-relativize.

    A .rels file at `xl/worksheets/_rels/sheet1.xml.rels` describes
    `xl/worksheets/sheet1.xml`, so its targets are relative to `xl/worksheets`.
    """
    base = posixpath.dirname(posixpath.dirname(rels_path))

    def one(element: bytes) -> bytes:
        if b'TargetMode="External"' in element:
            return element  # a URL, not a part in this package

        def target(m: re.Match) -> bytes:
            raw = m.group(1).decode("utf-8")
            part = raw[1:] if raw.startswith("/") else posixpath.normpath(
                posixpath.join(base, raw)
            )
            part = renames.get(part, part)
            rel = posixpath.relpath(part, base) if base else part
            return f'Target="{rel}"'.encode("utf-8")

        return REL_TARGET.sub(target, element)

    return REL_ELEMENT.sub(lambda m: one(m.group(0)), data)


def repack(path: Path) -> None:
    """Rewrite a saved workbook in place. See the note above."""
    with zipfile.ZipFile(path) as zin:
        entries = [(item.filename, zin.read(item.filename)) for item in zin.infolist()]

    renames = _rename_map([name for name, _ in entries])

    out = []
    for name, data in entries:
        if name.endswith(".rels"):
            data = _rewrite_rels(data, name, renames)
        elif name == "[Content_Types].xml":
            # Overrides name parts absolutely, with a leading slash.
            for old, new in renames.items():
                data = data.replace(
                    f'PartName="/{old}"'.encode("utf-8"),
                    f'PartName="/{new}"'.encode("utf-8"),
                )
        elif name == "docProps/core.xml":
            # openpyxl stamps `modified` with the wall clock as it saves, which
            # overwrites the fixed value set on the workbook a line earlier.
            data = MODIFIED.sub(
                rb"\g<1>" + FIXED_TIME.strftime("%Y-%m-%dT%H:%M:%SZ").encode() + rb"\g<2>",
                data,
            )
        out.append((renames.get(name, name), data))

    stamp = FIXED_TIME.timetuple()[:6]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in out:
            info = zipfile.ZipInfo(name, date_time=stamp)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0  # not "whichever OS regenerated it"
            zout.writestr(info, data)


def main():
    if not SAMPLES.is_dir():
        raise SystemExit(f"no samples/ directory at {SAMPLES}")

    paths = sorted(p for p in SAMPLES.glob("*.xlsx") if not p.name.startswith("~$"))
    if not paths:
        raise SystemExit("samples/ has no .xlsx files")

    books = {p.name: openpyxl.load_workbook(p) for p in paths}

    anon = Anonymizer()
    anon.harvest(books)

    # Clear the workbooks and nothing else. This used to rmtree the directory,
    # which took fixtures/README.md — the committed inventory of what each
    # fixture's edge cases are — with it every time anyone regenerated.
    FIXTURES.mkdir(exist_ok=True)
    for stale in FIXTURES.glob("*.xlsx"):
        stale.unlink()

    report = []
    for name, wb in books.items():
        cells = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    new = anon.scrub_cell(c.value)
                    if new != c.value:
                        c.value = new
                        cells += 1
            ws.title = anon.scrub_text(ws.title)

        # Strip authorship metadata — it names real people too.
        props = wb.properties
        props.creator = props.lastModifiedBy = "Royalty fixtures"
        props.title = props.subject = props.description = None
        # Fixed timestamps, so regenerating doesn't churn the diff.
        props.created = props.modified = FIXED_TIME

        # Filenames are kept verbatim — they name events and documents, not
        # people, and running them through the token map would rewrite words
        # like "Team" if one ever collided with a surname.
        out = FIXTURES / name
        wb.save(out)
        # Into the part layout Excel writes, so the app's reader can open it.
        repack(out)
        report.append((out.name, cells))

    print(f"{len(anon.first_map)} first names, {len(anon.last_map)} surnames, "
          f"{len(anon.phone_map)} phone numbers remapped\n")
    for fname, cells in report:
        print(f"  {fname:64s} {cells:5d} cells rewritten")


if __name__ == "__main__":
    main()
