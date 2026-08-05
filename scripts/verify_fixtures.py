#!/usr/bin/env python3
"""Prove that fixtures/ leaks nothing from samples/, and lost nothing either.

Run:  python3 scripts/verify_fixtures.py

Two questions, both of which have to be answered before fixtures/ can be
committed to a repository that will be public:

  1. Did any real name, phone number, or piece of document metadata survive
     into the fixtures?
  2. Did anonymizing quietly destroy the structure the fixtures exist to
     preserve — merged ranges, sheet geometry, the deliberate mess?

Exits non-zero on any leak, and prints a structural diff either way.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SAMPLES = ROOT / "samples"
FIXTURES = ROOT / "fixtures"

sys.path.insert(0, str(ROOT / "scripts"))
from anonymize_samples import (  # noqa: E402
    DIRECTION_MARKS,
    NOT_A_NAME,
    PERSON_RE,
    PHONE_RE,
    Anonymizer,
    strip_marks,
)


def cells(path: Path):
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    yield ws.title, c.coordinate, c.value
    wb.close()


def text_blobs(path: Path) -> list[str]:
    out = []
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        out.append(ws.title)
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    out.append(c.value)
                elif isinstance(c.value, (int, float)):
                    out.append(str(c.value))
    p = wb.properties
    out += [str(p.creator or ""), str(p.lastModifiedBy or ""), str(p.title or "")]
    wb.close()
    return out


def real_names(paths: list[Path]) -> set[str]:
    """Person-shaped strings in the originals, minus the known non-people.

    Deliberately re-derived here rather than taken from the anonymizer, so that
    a gap in the anonymizer's harvest shows up as a failure instead of being
    invisible to its own checker.
    """
    found = set()
    for p in paths:
        for blob in text_blobs(p):
            for m in PERSON_RE.finditer(blob):
                s = m.group(0).strip()
                toks = s.split()
                if any(t.isupper() and len(t) > 1 for t in toks):
                    continue  # ALL-CAPS free text is prose, not a name
                if any(t.lower().strip(".'’-") in NOT_A_NAME for t in toks):
                    continue
                found.add(s)
    return found


def real_phones(paths: list[Path]) -> set[str]:
    """Phone numbers as written, not as digit soup.

    Stripping all punctuation and then hunting 10-digit runs invents numbers
    that were never there — '9:45 AM - 12:45 PM' concatenates into one. Match
    the same shapes the anonymizer rewrites instead.
    """
    found = set()
    for p in paths:
        for blob in text_blobs(p):
            for m in PHONE_RE.finditer(strip_marks(blob)):
                found.add(m.group(2) + m.group(4) + m.group(6))
    return found


def structure(path: Path) -> dict:
    wb = openpyxl.load_workbook(path)
    s = {}
    for ws in wb.worksheets:
        nonempty = sum(
            1 for row in ws.iter_rows() for c in row if c.value is not None
        )
        s[ws.title] = (ws.max_row, ws.max_column, len(ws.merged_cells.ranges), nonempty)
    wb.close()
    return s


def main() -> int:
    sample_paths = sorted(p for p in SAMPLES.glob("*.xlsx") if not p.name.startswith("~$"))
    fixture_paths = sorted(p for p in FIXTURES.glob("*.xlsx"))

    if not fixture_paths:
        print("FAIL: fixtures/ is empty — run scripts/anonymize_samples.py")
        return 1

    print(f"checking {len(fixture_paths)} fixtures against {len(sample_paths)} samples\n")

    names = real_names(sample_paths)
    phones = real_phones(sample_paths)

    # The anonymizer's own vocabulary, checked token by token. This catches the
    # other failure mode: a name it knew about but failed to substitute.
    anon = Anonymizer()
    anon.harvest({p.name: openpyxl.load_workbook(p) for p in sample_paths})
    vocab = sorted(set(anon.first_map) | set(anon.last_map))

    # ALL-CAPS names hide from the Title-Case sweep. Find them the way a reader
    # would: a shouted run where one token is a name the anonymizer knows, so
    # its neighbours are names too. This is what caught 'ARIA DESAI' in the
    # orphan cell at Saturday!Q67, where the surname was replaced and the first
    # name was not.
    shouted = set()
    for p in sample_paths:
        for blob in text_blobs(p):
            for run in re.finditer(r"[A-Z][A-Z'’.-]{3,}(?:\s+[A-Z][A-Z'’.-]{3,})+", blob):
                toks = run.group(0).split()
                if any(t.lower() in vocab for t in toks):
                    shouted |= {t.lower() for t in toks
                                if t.isalpha() and t.lower() not in NOT_A_NAME}

    print(f"  {len(names)} person-shaped strings, {len(vocab)} name tokens, "
          f"{len(shouted)} shouted names, and {len(phones)} phone numbers "
          f"in the originals\n")

    failures = []

    # --- leak check -------------------------------------------------------
    for fp in fixture_paths:
        blobs = text_blobs(fp)
        joined = "\n".join(blobs)
        low = joined.lower()
        fixture_phones = {
            m.group(2) + m.group(4) + m.group(6)
            for m in PHONE_RE.finditer(strip_marks(joined))
        }

        for tok in vocab:
            if re.search(r"(?<![a-z])" + re.escape(tok) + r"(?![a-z])", low):
                failures.append(f"{fp.name}: real name token {tok!r} survived")

        for n in names:
            if re.search(r"(?<![A-Za-z])" + re.escape(n) + r"(?![A-Za-z])", joined):
                failures.append(f"{fp.name}: real name {n!r} survived")

        for n in sorted(shouted):
            if re.search(r"(?<![A-Za-z])" + re.escape(n) + r"(?![A-Za-z])", low):
                failures.append(f"{fp.name}: real ALL-CAPS name {n!r} survived")

        for ph in sorted(phones & fixture_phones):
            failures.append(f"{fp.name}: real phone {ph!r} survived")

    # --- structure check --------------------------------------------------
    print("  structure (sheets / merges / non-empty cells):")
    for sp, fp in zip(sample_paths, fixture_paths):
        a, b = structure(sp), structure(fp)
        if len(a) != len(b):
            failures.append(f"{fp.name}: sheet count {len(a)} -> {len(b)}")
        # Sheet names must survive verbatim, trailing spaces and all. They carry
        # meaning the importer depends on — team abbreviations especially, since
        # the roster stores team membership nowhere else.
        for x, y in zip(a, b):
            if x != y:
                failures.append(f"{fp.name}: sheet renamed {x!r} -> {y!r}")
        am = sum(v[2] for v in a.values())
        bm = sum(v[2] for v in b.values())
        an = sum(v[3] for v in a.values())
        bn = sum(v[3] for v in b.values())
        flag = "" if (am, an) == (bm, bn) else "   <-- CHANGED"
        print(f"    {fp.name[:52]:52s} {len(a):2d} sheets  "
              f"{am:4d} merges  {an:6d} cells{flag}")
        if am != bm:
            failures.append(f"{fp.name}: merged ranges {am} -> {bm}")
        if an != bn:
            failures.append(f"{fp.name}: non-empty cells {an} -> {bn}")

    # --- edge cases the fixtures exist to carry ---------------------------
    print("\n  edge cases:")
    roster = next(p for p in fixture_paths if p.name.startswith("Team Contact"))
    wb = openpyxl.load_workbook(roster)

    checks = []
    firsts_by_team = {}
    marks = floats = strings = padded = unicode_ph = blanks = 0
    for ws in wb.worksheets[:8]:
        names_seen = []
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            a, b, _diet, ph, _size = row
            if not (isinstance(a, str) and a.strip()):
                continue
            names_seen.append(f"{a.strip().rstrip('*').strip()} {b}")
            if "*" in a:
                marks += 1
            if a != a.strip():
                padded += 1
            # int vs float is an openpyxl round-trip detail; what the fixture
            # has to preserve is "stored as a number, not a string".
            if isinstance(ph, (int, float)) and not isinstance(ph, bool):
                floats += 1
            elif isinstance(ph, str):
                strings += 1
                if any(c in DIRECTION_MARKS for c in ph):
                    unicode_ph += 1
            elif ph is None:
                blanks += 1
        firsts_by_team[ws.title] = names_seen
    wb.close()

    dupes = sum(len(v) - len(set(v)) for v in firsts_by_team.values())
    checks.append(("asterisk food-restriction markers", marks, marks > 20))
    checks.append(("trailing-space names", padded, padded > 10))
    checks.append(("numeric-typed phones", floats, floats > 50))
    checks.append(("string-typed phones", strings, strings > 50))
    checks.append(("Unicode-direction-mark phones", unicode_ph, unicode_ph >= 1))
    checks.append(("dancers with no phone", blanks, blanks >= 1))
    checks.append(("within-team duplicate names", dupes, dupes >= 2))

    # Misspelling pairs: near-duplicate surnames on the day grids.
    task = next(p for p in fixture_paths if p.name.startswith("24-25"))
    wb = openpyxl.load_workbook(task)
    seen = set()
    for title in ("Thursday Schedule", "Friday Schedule", "Saturday Schedule ", "Sunday Schedule"):
        if title in wb.sheetnames:
            ws = wb[title]
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 2).value
                if isinstance(v, str) and v.strip():
                    seen.add(v.strip())
    wb.close()

    def dist(a, b):
        if abs(len(a) - len(b)) > 2:
            return 99
        prev = list(range(len(b) + 1))
        for i, ca in enumerate(a, 1):
            cur = [i]
            for j, cb in enumerate(b, 1):
                cur.append(min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + (ca != cb)))
            prev = cur
        return prev[-1]

    pairs = sum(
        1
        for i, a in enumerate(sorted(seen))
        for b in sorted(seen)[i + 1 :]
        if a != b and 0 < dist(a.lower(), b.lower()) <= 2
    )
    checks.append(("near-duplicate name spellings", pairs, pairs >= 4))

    for label, got, ok in checks:
        print(f"    {'ok  ' if ok else 'FAIL'} {label:38s} {got}")
        if not ok:
            failures.append(f"edge case lost: {label} ({got})")

    print()
    if failures:
        print(f"FAILED — {len(failures)} problem(s):")
        for f in failures[:40]:
            print(f"  - {f}")
        if len(failures) > 40:
            print(f"  ... and {len(failures) - 40} more")
        return 1

    print("PASSED — no real names, phones, or metadata in fixtures; "
          "structure and edge cases intact")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
